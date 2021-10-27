import json
import openpyxl
import toml
import logging
from openpyxl import load_workbook
from openpyxl import Workbook

logger = logging.getLogger(__name__)

class FileHandler:
   def __init__(self, file_object) -> None:
      self.file_object = file_object

   def open_json(self):
      try:
         with open(self.file_object) as f:
            data = json.load(f)
         return data
      except FileNotFoundError as e:
         logger.critical(f'Unable to find json file at {self.file_object}')

   def open_toml(self):
      try:
         with open(self.file_object) as f:
            data = toml.load(f)
         return data
      except FileNotFoundError as e:
         logger.critical(f'Unable to open toml file at {self.file_object}')

   def open_xlsx(self):
      try:
         wb = load_workbook(self.file_object) # Instantiate workbook
         return wb
      except:
         logger.critical(f'Unable to open toml file at {self.file_object}')
   
   def create_xlsx(self):
      try:
         wb = Workbook()
      except:
         logger.critical(f'Unable to open toml file at {self.file_object}')