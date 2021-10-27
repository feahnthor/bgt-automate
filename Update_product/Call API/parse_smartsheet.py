from openpyxl import Workbook
import toml
from dotmap import DotMap #keeping the imports here in case this module gets used later
import sys


sys.path += ['\\\\work\\tech\\henry\\programs\\python\\bgt-automate\\create_product'] # adds to path
from file_handler import FileHandler

class PasreSmartSheet:
  """
  
  """
  def __init__(self) -> None:
    self.conf = FileHandler('\\\\work\\tech\\Henry\\Programs\\Python\\bgt-automate\\Update_product\\Config\\.toml').open_toml()
    self.m_conf = DotMap(self.conf).var # now able to use dot notation for object/dict : means mapped_config
    self.wb = FileHandler('C:\\Users\\h.feahn\\Downloads\\Background Town 10-25-2021.xlsx').open_xlsx()

  def get_size(self):
    """
    Hex Color: #00ffed00
    wrong_RGB: (0, 255, 237)
    correct_RGB: (255, 237, 0)
    """
    self.ws = self.wb.active # opens first sheet
    # row = self.ws.rows
    foo = self.ws['B519']
    me = self.ws['A205']
    mes = me.fill.start_color.index
    _rgb = self.to_rgb(mes)
    col = foo.fill.start_color.index
    rgb = self.to_rgb(col)

    b = self.ws['B9:B790']
    a = self.ws['A9:A790']
    for name, size in zip(a,b):
      # print(size[0])
      if size[0].fill.start_color.index.lower() == '00ffed00' and size[0].value == 'yes':
        print(name[0].value)

    


    print(foo.value, col, rgb, foo.style, me.value, _rgb)


  def to_rgb(self, hex_color):
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4)) # This value is wrongs as the Code is incorrect
    return rgb
foo = PasreSmartSheet()
foo.get_size()
  